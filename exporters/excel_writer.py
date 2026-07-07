"""
Excel Writer — exports TestCase[] to .xlsx with proper formatting.

Column structure matches test-case-generate-with-prd-code standard:
A: 用例编号, B: 测试模块, C: 用例标题, D: 用例类型, E: 优先级,
F: 前置条件, G: 操作步骤, H: 预期结果, I: 用例来源, J: 备注
"""

from pathlib import Path
from typing import List

from schemas.models import TestCase


HEADERS = [
    "用例编号", "测试模块", "用例标题", "用例类型", "优先级",
    "前置条件", "操作步骤", "预期结果", "用例来源", "备注",
]


def write_excel(cases: List[TestCase], output_path: str) -> str:
    """Write test cases to an Excel file.

    Args:
        cases: Test cases to export.
        output_path: Path for the .xlsx file.

    Returns:
        The output path.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise ImportError("需要 openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # Header styling
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_align = Alignment(vertical="top", wrap_text=True)

    # Write headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, case in enumerate(cases, start=2):
        values = [
            case.id,
            case.module,
            case.title,
            case.case_type.value,
            case.priority.value,
            case.precondition,
            case.steps,
            case.expected,
            case.source,
            case.remark,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align
            cell.border = thin_border

    # Column widths
    col_widths = {
        "A": 16, "B": 14, "C": 40, "D": 10, "E": 8,
        "F": 28, "G": 40, "H": 40, "I": 30, "J": 30,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:J{len(cases) + 1}"

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"📊 Excel 已保存: {output_path}")
    return output_path
