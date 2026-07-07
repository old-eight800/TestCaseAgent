"""
Orchestrator — wires the full 5-stage pipeline: Collect → Extract → Design → Review → Fix.

Mirrors agentscope-testgen's Orchestrator pattern:
  Phase 1: Input collection (static)
  Phase 2: Rule extraction (LLM)
  Phase 3: Case design (LLM)
  Phase 4: Review (rule-engine + LLM)
  Phase 5: Fix loop (up to MAX_ROUNDS)
  Phase 6: Export
"""

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional, List

from agents.llm_client import LLMClient
from agents.input_collector import InputCollector
from agents.rule_extractor import RuleExtractor
from agents.case_designer import CaseDesigner
from agents.reviewer import Reviewer
from agents.fixer import Fixer
from exporters.excel_writer import write_excel
from exporters.xmind_writer import write_xmind
from schemas.models import RawInputs, Rule, TestCase, ReviewReport


class Orchestrator:
    """Central pipeline orchestrator for test case generation."""

    MAX_REVIEW_ROUNDS = 3

    def __init__(self, client: Optional[LLMClient] = None):
        self.client = client or LLMClient()
        self.collector = InputCollector()
        self.extractor = RuleExtractor(self.client)
        self.designer = CaseDesigner(self.client)
        self.reviewer = Reviewer(self.client)
        self.fixer = Fixer(self.client)

    def run(self,
            module: str,
            prd_path: Optional[str] = None,
            code_paths: Optional[List[str]] = None,
            design_path: Optional[str] = None,
            few_shot_paths: Optional[List[str]] = None,
            output_dir: Optional[str] = None,
            formats: str = "xlsx,xmind",
            max_rounds: int = 3,
            force_sources: str = "all",
            context_dir: Optional[str] = None,
            no_review: bool = False,
            ) -> dict:
        """Execute the full pipeline.

        Returns a summary dict for CLI display.
        """
        ts = datetime.now().strftime("%Y%m%d_%H%M")

        # ── Phase 1: Collect inputs ───────────────────────────────
        print("=" * 60)
        print("🚀 TestCase Agent Pipeline 启动")
        print(f"   Module: {module}")
        print(f"   PRD: {prd_path or '(无)'}")
        print(f"   Code: {len(code_paths or [])} 个文件")
        print(f"   Design: {design_path or '(无)'}")
        print("=" * 60)

        print("\n📋 Phase 1: 输入收集...")
        raw = self.collector.collect(
            prd_path=prd_path,
            code_paths=code_paths,
            design_path=design_path,
            module=module,
            context_dir=context_dir,
        )
        if not raw.has_any():
            print("❌ 无任何有效输入，退出")
            return {"status": "failed", "reason": "no_inputs"}

        # ── Phase 2: Extract rules ────────────────────────────────
        print(f"\n📐 Phase 2: 规则抽取...")
        rules = self.extractor.extract(raw, force_sources=force_sources)
        if not rules:
            print("❌ 未抽取到任何规则，退出")
            return {"status": "failed", "reason": "no_rules"}

        print(f"   总规则数: {len(rules)}")
        for r in rules:
            diff_tag = f" {r.diff_mark.value}" if r.diff_mark.value else ""
            print(f"     [{r.source.value}] {r.id}: {r.description[:60]}{diff_tag}")

        # Save intermediate
        self._save_intermediate(rules, output_dir, f"rules_{module}_{ts}.json")

        # ── Phase 3: Design cases ─────────────────────────────────
        print(f"\n📝 Phase 3: 用例设计...")
        few_shot_text = self._load_few_shot(few_shot_paths)
        cases = self.designer.design(rules, few_shot_text=few_shot_text)
        if not cases:
            print("❌ 未生成任何用例，退出")
            return {"status": "failed", "reason": "no_cases"}
        print(f"   总用例数: {len(cases)}")

        # Save intermediate
        self._save_intermediate(cases, output_dir, f"candidate_cases_{module}_{ts}.json")

        # ── Phase 4-5: Review → Fix loop ──────────────────────────
        if no_review:
            print("\n⏭️  跳过审查（--no-review）")
            final_cases = cases
            review_report = ReviewReport(score=100)
            rounds = 0
        else:
            final_cases, review_report, rounds = self._review_loop(
                cases, rules, raw, max_rounds
            )

        # ── Phase 6: Export ───────────────────────────────────────
        print(f"\n📦 Phase 6: 导出...")
        output_dir = Path(output_dir) if output_dir else (
            Path(__file__).parent.parent / "output" / module
        )
        output_dir.mkdir(parents=True, exist_ok=True)

        results = {}
        for fmt in formats.split(","):
            fmt = fmt.strip().lower()
            if fmt == "xlsx":
                out_path = output_dir / f"test_cases_{module}_{ts}.xlsx"
                write_excel(final_cases, str(out_path))
                results["xlsx"] = str(out_path)
            elif fmt == "xmind":
                out_path = output_dir / f"test_cases_{module}_{ts}.xmind"
                write_xmind(final_cases, str(out_path))
                results["xmind"] = str(out_path)

        # Save final review report
        self._save_intermediate(
            {"score": review_report.score,
             "total_issues": review_report.total_issues,
             "critical": review_report.critical,
             "warnings": review_report.warnings,
             "issues": [{"severity": i.severity, "dimension": i.dimension,
                         "location": i.location, "description": i.description}
                        for i in review_report.issues]},
            output_dir, f"review_report_{module}_{ts}.json"
        )

        return {
            "status": "success" if review_report.is_passing() else "warning",
            "module": module,
            "rules_count": len(rules),
            "cases_count": len(final_cases),
            "output_files": results,
            "review_score": review_report.score,
            "review_rounds": rounds,
        }

    def _review_loop(self, cases: List[TestCase], rules: List[Rule],
                     raw: RawInputs, max_rounds: int) -> tuple:
        """Run the review-fix loop up to max_rounds."""
        current_cases = cases

        for round_num in range(1, max_rounds + 1):
            print(f"\n🔍 Phase 4.{round_num}: 审查轮次 {round_num}/{max_rounds}...")

            # Rule engine always runs; LLM review only on first and last rounds
            use_llm = (round_num == 1 or round_num == max_rounds)
            report = self.reviewer.review(current_cases, rules,
                                          scope_text=raw.scope_md,
                                          use_llm=use_llm)

            print(f"   Score: {report.score}/100")
            print(f"   Critical: {report.critical}, Warnings: {report.warnings}")

            if report.is_passing():
                print("   ✅ 审查通过！")
                return current_cases, report, round_num

            if round_num < max_rounds:
                print(f"   🔄 修复中...")
                current_cases = self.fixer.fix(report, current_cases, rules)

        # Last round — use whatever we have
        return current_cases, report, round_num

    @staticmethod
    def _load_few_shot(paths: Optional[List[str]]) -> Optional[str]:
        """Load few-shot examples from xlsx/xmind paths as style reference."""
        if not paths:
            return None

        parts = []
        for p in paths:
            path = Path(p)
            if not path.exists():
                continue
            suffix = path.suffix.lower()
            try:
                if suffix in (".xlsx", ".xls"):
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(str(path), read_only=True)
                        ws = wb.active
                        rows = []
                        for row in ws.iter_rows(values_only=True, max_row=6):
                            rows.append(" | ".join(str(c) if c else "" for c in row))
                        parts.append(f"### {path.name} (前 5 行)\n" + "\n".join(rows))
                    except ImportError:
                        parts.append(f"### {path.name}\n(需要 openpyxl 读取)")
                elif suffix == ".xmind":
                    import zipfile
                    with zipfile.ZipFile(str(path)) as zf:
                        content = json.loads(zf.read("content.json"))
                        # Extract top-level structure only
                        parts.append(f"### {path.name}\n" +
                                     json.dumps(content, ensure_ascii=False, indent=2)[:2000])
            except Exception as e:
                print(f"⚠️  加载 Few-shot 失败: {p} — {e}")

        return "\n\n".join(parts) if parts else None

    @staticmethod
    def _save_intermediate(data, output_dir: Optional[str], filename: str):
        """Save intermediate results as JSON for auditability."""
        if output_dir:
            base = Path(output_dir)
        else:
            base = Path(__file__).parent.parent / "output"
        base.mkdir(parents=True, exist_ok=True)

        out_path = base / filename

        if isinstance(data, list):
            # Convert dataclasses to dicts
            if data and hasattr(data[0], '__dataclass_fields__'):
                from dataclasses import asdict
                data = [asdict(d) for d in data]

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

        print(f"   💾 中间产物已保存: {out_path}")


# ── Convenience entry ─────────────────────────────────────────────

def run(module: str, **kwargs) -> dict:
    """Convenience entry point for programmatic use."""
    orch = Orchestrator()
    return orch.run(module=module, **kwargs)
